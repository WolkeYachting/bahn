"""Minimal Excel export.

Two sheets:
  Vorfälle    - one row per qualifying incident (the main output)
  Diagnose    - trips that never appeared in live data (sanity check)

If you want, you can ignore the Diagnose sheet entirely. It's there to make
sure no relevant trips quietly slipped through the cracks.
"""

import io
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from storage import Storage

BASE = Path(__file__).parent
INCIDENTS_FILE = "incidents.json"
EXPORT_FILE = "exports/bahn_verspaetungen.xlsx"

_storage = Storage()

HEADER_FILL = PatternFill("solid", start_color="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
DEFAULT_FONT = Font(name="Arial")

FILL_BY_CATEGORY = {
    "full_cancellation": PatternFill("solid", start_color="FFCCCC"),
    "partial_cancellation": PatternFill("solid", start_color="FCE4D6"),
    "delayed": None,
}

CATEGORY_LABEL = {
    "delayed": "Verspätung",
    "partial_cancellation": "Teilausfall",
    "full_cancellation": "Vollausfall",
}

COLUMNS = [
    ("Datum", "date"),
    ("Zug", "line"),
    ("Richtung", "direction"),
    ("Startbahnhof", "start_station"),
    ("Soll-Abfahrt", "start_planned_departure"),
    ("Zielbahnhof", "end_station"),
    ("Soll-Ankunft", "end_planned_arrival"),
    ("Ist-Ankunft", "end_actual_arrival"),
    ("Verspätung (min)", "delay_minutes"),
    ("Kategorie", "category_label"),
    ("Ausfall ab", "cancelled_from"),
    ("Erzwungen abgeschlossen", "force_finished"),
    ("Hinweis", "note"),
]

DIAG_COLUMNS = [
    ("Datum", "date"),
    ("Zug", "line"),
    ("Richtung", "direction"),
    ("Trip-ID", "trip_id"),
    ("Grund", "reason"),
]


def fmt_iso(iso: str | None) -> str:
    if not iso:
        return ""
    try:
        return datetime.fromisoformat(iso.replace("Z", "+00:00")).strftime("%d.%m.%Y %H:%M")
    except Exception:
        return iso


def _write_sheet(ws, columns, rows, fill_fn=None):
    for j, (label, _) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=j, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    for r, row in enumerate(rows, start=2):
        fill = fill_fn(row) if fill_fn else None
        for j, (_, key) in enumerate(columns, start=1):
            value = row.get(key)
            if value is None:
                value = ""
            cell = ws.cell(row=r, column=j, value=value)
            cell.font = DEFAULT_FONT
            if fill is not None:
                cell.fill = fill

    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        longest = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), 45)


def generate(output_path: Path | str | None = None) -> Path:
    raw = _storage.read(INCIDENTS_FILE)
    store = json.loads(raw) if raw else {"incidents": [], "diagnostics": []}
    incidents = store.get("incidents", [])
    diagnostics = store.get("diagnostics", [])

    # Prepare display rows for incidents
    inc_rows = []
    for inc in incidents:
        view = dict(inc)
        view["category_label"] = CATEGORY_LABEL.get(
            inc.get("category"), inc.get("category", ""))
        view["start_planned_departure"] = fmt_iso(inc.get("start_planned_departure"))
        view["end_planned_arrival"] = fmt_iso(inc.get("end_planned_arrival"))
        view["end_actual_arrival"] = fmt_iso(inc.get("end_actual_arrival"))
        view["force_finished"] = "ja" if inc.get("force_finished") else ""
        inc_rows.append(view)

    wb = Workbook()
    ws = wb.active
    ws.title = "Vorfälle"
    _write_sheet(ws, COLUMNS, inc_rows,
                 fill_fn=lambda r: FILL_BY_CATEGORY.get(r.get("category")))

    ws = wb.create_sheet("Diagnose")
    _write_sheet(ws, DIAG_COLUMNS, diagnostics)

    if output_path is not None:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        print(f"Excel geschrieben: {output_path} "
              f"({len(incidents)} Vorfälle, {len(diagnostics)} Diagnose)")
        return output_path

    buf = io.BytesIO()
    wb.save(buf)
    _storage.write_binary(
        EXPORT_FILE,
        buf.getvalue(),
        commit_message=(
            f"export: {len(incidents)} Vorfälle, {len(diagnostics)} Diagnose"
        ),
    )
    final = BASE / EXPORT_FILE
    print(f"Excel geschrieben: {final} "
          f"({len(incidents)} Vorfälle, {len(diagnostics)} Diagnose)")
    return final


if __name__ == "__main__":
    generate()
