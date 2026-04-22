"""Tests for the full pipeline.

Covers:
- timeutil.logical_day_for boundary semantics
- poll_day state machine (active → finished transition with grace)
- analyze_day classification (delayed / partial / full / on-time / scheduled)
- analyze_day force-finish behaviour
- Excel generation
"""

import json
import os
from datetime import datetime, timedelta, timezone
from pathlib import Path
from unittest.mock import patch

import timeutil
import poll_day
import analyze_day
import generate_excel


STATION_ID = "8000148"
STATION_NAME = "Paderborn Hbf"


# ---- helpers --------------------------------------------------------------

def stopover(stop_id, name, planned_arr=None, actual_arr=None,
             delay_min=None, planned_dep=None, actual_dep=None,
             cancelled=False):
    return {
        "stop_id": stop_id,
        "stop_name": name,
        "planned_arrival": planned_arr,
        "arrival": actual_arr,
        "arrival_delay_minutes": delay_min,
        "planned_departure": planned_dep or planned_arr,
        "departure": actual_dep or actual_arr,
        "departure_delay_minutes": delay_min,
        "cancelled": cancelled,
        "arrival_platform": None,
        "planned_arrival_platform": None,
    }


def trip_data(trip_id, line, direction, stops, cancelled=False):
    """Build a trip 'data' dict as it would live inside a day-log entry."""
    idx = next((i for i, s in enumerate(stops) if s["stop_id"] == STATION_ID), 0)
    return {
        "trip_id": trip_id,
        "line_name": line,
        "product": "regional",
        "origin": "Paderborn Hbf",
        "destination": stops[-1]["stop_name"],
        "direction": direction,
        "cancelled": cancelled,
        "station_stopover_index": idx,
        "stopovers": stops,
    }


def trip_entry(data, status="finished", force_finished=False):
    return {
        "data": data,
        "status": status,
        "first_seen_at": "2026-04-21T05:00:00+00:00",
        "last_seen_at": "2026-04-21T08:00:00+00:00",
        "finished_at": "2026-04-21T08:30:00+00:00" if status == "finished" else None,
        "force_finished": force_finished,
    }


# ---- 1. timeutil.logical_day_for ------------------------------------------

def test_logical_day_boundary():
    # 02:00 Berlin on April 22 → still belongs to logical day "April 21"
    dt = datetime(2026, 4, 22, 2, 0, tzinfo=timeutil.BERLIN)
    assert timeutil.logical_day_for(dt, 3, 30) == "2026-04-21"

    # 03:29 Berlin → still April 21
    dt = datetime(2026, 4, 22, 3, 29, tzinfo=timeutil.BERLIN)
    assert timeutil.logical_day_for(dt, 3, 30) == "2026-04-21"

    # 03:30 Berlin → flips to April 22
    dt = datetime(2026, 4, 22, 3, 30, tzinfo=timeutil.BERLIN)
    assert timeutil.logical_day_for(dt, 3, 30) == "2026-04-22"

    # Late evening → today
    dt = datetime(2026, 4, 22, 23, 0, tzinfo=timeutil.BERLIN)
    assert timeutil.logical_day_for(dt, 3, 30) == "2026-04-22"
    print("[ok] logical_day_for handles 03:30 boundary correctly")


# ---- 2. poll_day state-machine: trip transitions to 'finished' ------------

def test_transition_to_finished_after_grace():
    # Planned arrival at destination 10 minutes ago, grace 30 min → not yet
    now = datetime.now(timezone.utc)
    arr_iso = (now - timedelta(minutes=10)).isoformat()
    data = trip_data(
        "x", "RE 1", "Köln",
        [stopover(STATION_ID, "Paderborn Hbf",
                  planned_dep=(now - timedelta(hours=1)).isoformat()),
         stopover("X", "Hamm", planned_arr=arr_iso)],
    )
    assert poll_day.transition_to_finished(data, now, 30) is False
    print("[ok] still-active trip 10 min past arrival NOT finished (grace=30)")

    # 31 min past arrival → finish
    arr_iso = (now - timedelta(minutes=31)).isoformat()
    data = trip_data(
        "x", "RE 1", "Köln",
        [stopover(STATION_ID, "Paderborn Hbf",
                  planned_dep=(now - timedelta(hours=2)).isoformat()),
         stopover("X", "Hamm", planned_arr=arr_iso)],
    )
    assert poll_day.transition_to_finished(data, now, 30) is True
    print("[ok] still-active trip 31 min past arrival IS finished (grace=30)")


# ---- 3. analyze_day classification ----------------------------------------

def test_classify_on_time():
    data = trip_data("t", "RE 11", "Kassel", [
        stopover(STATION_ID, "Paderborn Hbf",
                 "2026-04-21T14:20:00+02:00", "2026-04-21T14:20:00+02:00", 0),
        stopover("X", "Altenbeken",
                 "2026-04-21T14:35:00+02:00", "2026-04-21T14:40:00+02:00", 5),
        stopover("Y", "Kassel",
                 "2026-04-21T15:30:00+02:00", "2026-04-21T15:38:00+02:00", 8),
    ])
    assert analyze_day.classify_trip(data, 60) is None
    print("[ok] on-time trip not classified as incident")


def test_classify_confirmed_delay():
    data = trip_data("t", "RE 11", "Kassel", [
        stopover(STATION_ID, "Paderborn Hbf",
                 "2026-04-21T14:20:00+02:00", "2026-04-21T14:20:00+02:00", 0),
        stopover("X", "Warburg",
                 "2026-04-21T14:45:00+02:00", "2026-04-21T15:50:00+02:00", 65),
        stopover("Y", "Kassel",
                 "2026-04-21T15:30:00+02:00", "2026-04-21T16:20:00+02:00", 50),
    ])
    inc = analyze_day.classify_trip(data, 60)
    assert inc is not None
    assert inc["category"] == "delayed"
    assert inc["end_station"] == "Warburg"
    assert inc["delay_minutes"] == 65
    print("[ok] confirmed +65 at Warburg recorded as delay (Warburg is target)")


def test_classify_prognosis_only_skipped():
    # +70 predicted at Warburg but actual arrival never set → just prognosis
    data = trip_data("t", "RB 89", "Kassel", [
        stopover(STATION_ID, "Paderborn Hbf",
                 "2026-04-21T16:00:00+02:00", "2026-04-21T16:00:00+02:00", 0),
        stopover("X", "Warburg",
                 "2026-04-21T16:30:00+02:00", None, 70),
        stopover("Y", "Kassel",
                 "2026-04-21T17:10:00+02:00", "2026-04-21T17:55:00+02:00", 45),
    ])
    inc = analyze_day.classify_trip(data, 60)
    assert inc is None, f"prognosis should not count: {inc}"
    print("[ok] prognosis-only delay not recorded")


def test_classify_partial_cancellation():
    data = trip_data("t", "RB 74", "Bielefeld", [
        stopover(STATION_ID, "Paderborn Hbf",
                 "2026-04-21T10:00:00+02:00", "2026-04-21T10:00:00+02:00", 0),
        stopover("X", "Schloß Neuhaus",
                 "2026-04-21T10:10:00+02:00", "2026-04-21T10:10:00+02:00", 0),
        stopover("Y", "Bielefeld",
                 "2026-04-21T10:45:00+02:00", None, cancelled=True),
    ])
    inc = analyze_day.classify_trip(data, 60)
    assert inc is not None
    assert inc["category"] == "partial_cancellation"
    assert inc["cancelled_from"] == "Bielefeld"
    assert inc["end_station"] == "Schloß Neuhaus"
    print("[ok] partial cancellation captured with last reached stop")


def test_classify_full_cancellation():
    data = trip_data("t", "S 5", "Hameln", [
        stopover(STATION_ID, "Paderborn Hbf",
                 "2026-04-21T08:00:00+02:00", None, cancelled=True),
        stopover("Z", "Altenbeken",
                 "2026-04-21T08:15:00+02:00", None, cancelled=True),
    ], cancelled=True)
    inc = analyze_day.classify_trip(data, 60)
    assert inc is not None
    assert inc["category"] == "full_cancellation"
    assert inc["cancelled_from"] == "Paderborn Hbf"
    print("[ok] full cancellation captured")


# ---- 4. analyze_day force-finish ------------------------------------------

def test_force_finish_marks_active_trips():
    log = {
        "date": "2026-04-21",
        "trips": {
            "a": {"status": "active",
                  "data": trip_data("a", "RE 1", "Köln", [
                      stopover(STATION_ID, "Paderborn Hbf",
                               "2026-04-21T10:00:00+02:00",
                               "2026-04-21T10:00:00+02:00", 0)
                  ])},
            "b": {"status": "finished",
                  "data": trip_data("b", "RE 2", "Soest", [
                      stopover(STATION_ID, "Paderborn Hbf",
                               "2026-04-21T11:00:00+02:00",
                               "2026-04-21T11:00:00+02:00", 0)
                  ])},
        },
    }
    forced = analyze_day.force_finish_active_trips(log, "2026-04-22T07:00:00+00:00")
    assert forced == 1
    assert log["trips"]["a"]["status"] == "finished"
    assert log["trips"]["a"]["force_finished"] is True
    assert log["trips"]["b"].get("force_finished") is None
    print("[ok] force-finish flips active → finished and sets marker")


# ---- 5. End-to-end with synthetic day-log ---------------------------------

def test_end_to_end(tmp_path):
    os.environ.pop("GITHUB_TOKEN", None)
    os.environ.pop("GITHUB_REPO", None)

    # Redirect storage to tmp dir
    poll_day._storage.local_base = tmp_path
    analyze_day._storage.local_base = tmp_path
    generate_excel._storage.local_base = tmp_path

    try:
        # Build a synthetic day-log with mixed trip types
        date = "2026-04-21"
        log = {
            "date": date,
            "closed": False,
            "station_id": STATION_ID,
            "station_name": STATION_NAME,
            "trips": {
                "ontime": trip_entry(trip_data("ontime", "RE 11", "Kassel", [
                    stopover(STATION_ID, "Paderborn Hbf",
                             "2026-04-21T14:20:00+02:00",
                             "2026-04-21T14:20:00+02:00", 0),
                    stopover("X", "Altenbeken",
                             "2026-04-21T14:35:00+02:00",
                             "2026-04-21T14:38:00+02:00", 3),
                ])),
                "delay": trip_entry(trip_data("delay", "RE 11", "Kassel", [
                    stopover(STATION_ID, "Paderborn Hbf",
                             "2026-04-21T15:00:00+02:00",
                             "2026-04-21T15:00:00+02:00", 0),
                    stopover("Y", "Warburg",
                             "2026-04-21T15:25:00+02:00",
                             "2026-04-21T16:30:00+02:00", 65),
                ])),
                "scheduled": {
                    "data": trip_data("scheduled", "S 5", "Hameln", [
                        stopover(STATION_ID, "Paderborn Hbf",
                                 "2026-04-21T22:00:00+02:00", None, None)
                    ]),
                    "status": "scheduled",
                    "first_seen_at": None,
                    "last_seen_at": None,
                    "force_finished": False,
                },
                "stuck_active": {
                    # active but planned arrival was hours ago — analyze() will
                    # force-finish this and then classify
                    "data": trip_data("stuck_active", "RB 89", "Kassel", [
                        stopover(STATION_ID, "Paderborn Hbf",
                                 "2026-04-21T16:00:00+02:00",
                                 "2026-04-21T16:00:00+02:00", 0),
                        stopover("Z", "Kassel",
                                 "2026-04-21T17:10:00+02:00",
                                 "2026-04-21T18:25:00+02:00", 75),
                    ]),
                    "status": "active",
                    "first_seen_at": "2026-04-21T15:00:00+00:00",
                    "last_seen_at": "2026-04-21T17:00:00+00:00",
                    "force_finished": False,
                },
            },
            "polls": [],
        }
        (tmp_path / "logs").mkdir(exist_ok=True)
        (tmp_path / "logs" / f"{date}.json").write_text(
            json.dumps(log), encoding="utf-8"
        )

        summary = analyze_day.run(date)
        assert summary["incidents_added"] == 2, summary
        assert summary["diagnostics_added"] == 1, summary  # the 'scheduled' one
        assert summary["force_finished_count"] == 1, summary

        # Excel generation
        out = generate_excel.generate(tmp_path / "out.xlsx")
        from openpyxl import load_workbook
        wb = load_workbook(out)
        assert wb.sheetnames == ["Vorfälle", "Diagnose"], wb.sheetnames
        ws = wb["Vorfälle"]
        assert ws.max_row == 3, f"expected 1 header + 2 incidents, got {ws.max_row}"
        ws = wb["Diagnose"]
        assert ws.max_row == 2, f"expected 1 header + 1 diagnostic, got {ws.max_row}"

        # Day-log should now be marked closed
        log_after = json.loads(
            (tmp_path / "logs" / f"{date}.json").read_text(encoding="utf-8")
        )
        assert log_after["closed"] is True
        assert log_after["trips"]["stuck_active"]["force_finished"] is True

        # Force-finished flag should propagate to incidents
        store = json.loads(
            (tmp_path / "incidents.json").read_text(encoding="utf-8")
        )
        forced_inc = next(i for i in store["incidents"]
                          if i["trip_id"] == "stuck_active")
        assert forced_inc["force_finished"] is True

        # Re-running analysis on the same date should be a no-op
        summary2 = analyze_day.run(date)
        assert summary2.get("skipped") is True
    finally:
        poll_day._storage.local_base = poll_day.BASE
        analyze_day._storage.local_base = analyze_day.BASE
        generate_excel._storage.local_base = generate_excel.BASE
    print("[ok] end-to-end: 4 trips → 2 incidents, 1 diagnostic, 1 force-finished")


# ---- 6. poll_day integration with mocked HTTP -----------------------------

def test_poll_integration_with_mocks(tmp_path):
    os.environ.pop("GITHUB_TOKEN", None)
    os.environ.pop("GITHUB_REPO", None)
    poll_day._storage.local_base = tmp_path

    now = datetime.now(timezone.utc)
    plan_dep = (now + timedelta(hours=1)).isoformat()
    plan_arr = (now + timedelta(hours=2)).isoformat()

    fake_departures = [
        {"tripId": "trip-1", "line": {"name": "RE 11", "product": "regional"}},
    ]

    fake_trip = {
        "id": "trip-1",
        "line": {"name": "RE 11", "product": "regional"},
        "origin": {"name": "Paderborn Hbf"},
        "destination": {"name": "Kassel"},
        "direction": "Kassel",
        "cancelled": False,
        "stopovers": [
            {
                "stop": {"id": STATION_ID, "name": "Paderborn Hbf"},
                "plannedDeparture": plan_dep,
                "departure": plan_dep,
                "departureDelay": 0,
            },
            {
                "stop": {"id": "X", "name": "Kassel"},
                "plannedArrival": plan_arr,
                "arrival": plan_arr,
                "arrivalDelay": 0,
            },
        ],
    }

    try:
        with patch.object(poll_day, "fetch_departures",
                          return_value=fake_departures), \
             patch.object(poll_day, "fetch_trip", return_value=fake_trip), \
             patch.object(poll_day.time, "sleep", lambda *a: None):
            stats = poll_day.poll_once()

        assert stats["trips_fetched"] >= 1, stats
        # Day log should exist with one active trip — find it by scanning files
        logs_dir = tmp_path / "logs"
        log_files = list(logs_dir.glob("*.json"))
        assert log_files, "no day log was written"
        found = False
        for f in log_files:
            log = json.loads(f.read_text(encoding="utf-8"))
            if "trip-1" in log.get("trips", {}):
                assert log["trips"]["trip-1"]["status"] == "active"
                found = True
                break
        assert found, f"trip-1 not found in any log file: {[f.name for f in log_files]}"
    finally:
        poll_day._storage.local_base = poll_day.BASE
    print("[ok] poll_day integration: trip seen → added as 'active' to correct day log")


# ---- runner ---------------------------------------------------------------

if __name__ == "__main__":
    import tempfile
    test_logical_day_boundary()
    test_transition_to_finished_after_grace()
    test_classify_on_time()
    test_classify_confirmed_delay()
    test_classify_prognosis_only_skipped()
    test_classify_partial_cancellation()
    test_classify_full_cancellation()
    test_force_finish_marks_active_trips()
    with tempfile.TemporaryDirectory() as d:
        test_end_to_end(Path(d))
    with tempfile.TemporaryDirectory() as d:
        test_poll_integration_with_mocks(Path(d))
    print("\nAlle Tests grün.")
